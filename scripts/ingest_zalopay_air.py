#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from dotenv import load_dotenv
except ImportError:  # pragma: no cover - optional in non-app environments
    load_dotenv = None


ROOT = Path(__file__).resolve().parents[1]
ZALOPAY_DIR = ROOT / "zalopay"
DATA_DIR = ROOT / "data"
DOMAIN = "Zalopay AIR/OTA"
OWNER = "Zalopay OTA"
SOURCE_EVENT_ID = "zlp_air_reference_import"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def slug(value: str) -> str:
    cleaned = normalize_text(value).lower()
    cleaned = re.sub(r"[^a-z0-9]+", "_", cleaned)
    return cleaned.strip("_")


def unique(values: list[Any]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        cleaned = normalize_text(value)
        key = cleaned.casefold()
        if cleaned and key not in seen:
            seen.add(key)
            result.append(cleaned)
    return result


def load_json(path: Path, fallback: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        return fallback
    with path.open(encoding="utf-8") as handle:
        return json.load(handle)


def save_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(data, handle, ensure_ascii=False, indent=2, sort_keys=True)
        handle.write("\n")


def strip_markdown(value: str) -> str:
    text = value.replace("**", "")
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"\[\[([^\]]+)\]\]", r"\1", text)
    return normalize_text(text)


def extract_aliases(block: str) -> list[str]:
    aliases: list[str] = []
    for match in re.finditer(r"\*Alias:\*\s*(.+)", block):
        raw = match.group(1)
        aliases.extend(part.strip(" \"'`.;") for part in raw.split(","))
    return unique(aliases)


def infer_kind(name: str, definition: str) -> str:
    text = f"{name} {definition}".lower()
    if any(marker in text for marker in ["rate", "ratio", "tpv", "user", "aov", "arppu", "atp", "pu", "rr"]):
        return "metric"
    if name.lower() in {"route", "booking_window", "ticket_num", "tracking_session_id"}:
        return "dimension"
    return "term"


def parse_metric_definitions(path: Path) -> dict[str, dict[str, Any]]:
    text = path.read_text(encoding="utf-8")
    matches = list(re.finditer(r"^(?:-\s*)?\*\*(.+?)\*\*\s*=\s*(.+)$", text, flags=re.MULTILINE))
    records: dict[str, dict[str, Any]] = {}
    timestamp = now_iso()

    for index, match in enumerate(matches):
        raw_name = strip_markdown(match.group(1))
        first_definition = strip_markdown(match.group(2))
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[match.end() : end]
        name = re.split(r"\s*\(", raw_name, maxsplit=1)[0].strip()
        metric_id = f"kn_zlp_air_{slug(name)}"
        aliases = extract_aliases(block)
        if raw_name != name:
            aliases.append(raw_name)
        conditions = []
        for line in block.splitlines():
            cleaned = strip_markdown(line.lstrip("- ").strip())
            if cleaned and not cleaned.lower().startswith("alias:"):
                conditions.append(cleaned)
        records[metric_id] = {
            "id": metric_id,
            "kind": infer_kind(name, first_definition),
            "name": name,
            "canonical_definition": first_definition,
            "logic": "Zalopay AIR/OTA approved metric definition. Override generic definitions for Air/OTA requests.",
            "examples": [],
            "paraphrases": unique(aliases),
            "formula": "",
            "conditions": unique(conditions),
            "domain": DOMAIN,
            "owner": OWNER,
            "created_by": "zalopay_air_ingest",
            "confidence": 0.98,
            "version": 1,
            "status": "approved",
            "evidence_event_ids": [SOURCE_EVENT_ID],
            "candidate_ids": [],
            "change_history": [],
            "created_at": timestamp,
            "updated_at": timestamp,
        }
    return records


def parse_markdown_table_rows(path: Path) -> dict[str, dict[str, str]]:
    rows: dict[str, dict[str, str]] = {}
    current_headers: list[str] = []
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not (line.startswith("|") and line.endswith("|")):
            continue
        cells = [strip_markdown(cell.strip()) for cell in line.strip("|").split("|")]
        if cells and cells[0] == "---":
            continue
        if "#" in cells and ("Câu hỏi" in cells or "Spec" in cells):
            current_headers = cells
            continue
        if not current_headers or not cells or not re.fullmatch(r"\d+", cells[0]):
            continue
        item = {current_headers[i]: cells[i] for i in range(min(len(current_headers), len(cells)))}
        rows[cells[0]] = item
    return rows


def parse_playbook(path: Path) -> dict[str, dict[str, str]]:
    text = path.read_text(encoding="utf-8")
    pattern = re.compile(r"^###\s+((?:#\d+\s*(?:/\s*)?)+)\.\s*(.+)$", flags=re.MULTILINE)
    matches = list(pattern.finditer(text))
    sections: dict[str, dict[str, str]] = {}
    for index, match in enumerate(matches):
        numbers = re.findall(r"#(\d+)", match.group(1))
        title = strip_markdown(match.group(2))
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        block = text[match.end() : end].strip()
        sql_match = re.search(r"```sql\s*(.*?)```", block, flags=re.DOTALL | re.IGNORECASE)
        make_clear_match = re.search(r"\*\*Make clear:\*\*\s*(.+)", block)
        for number in numbers:
            sections[number] = {
                "title": title,
                "make_clear": strip_markdown(make_clear_match.group(1)) if make_clear_match else "",
                "sql": sql_match.group(1).strip() if sql_match else "",
                "block": block,
            }
    return sections


def extract_tables(sql: str) -> list[str]:
    return unique(re.findall(r"\b(?:FROM|JOIN)\s+([a-zA-Z_][a-zA-Z0-9_]*)", sql, flags=re.IGNORECASE))


def parse_question_examples(catalog_path: Path, playbook_path: Path) -> dict[str, dict[str, Any]]:
    catalog = parse_markdown_table_rows(catalog_path)
    playbook = parse_playbook(playbook_path)
    examples: dict[str, dict[str, Any]] = {}
    timestamp = now_iso()
    for number, section in sorted(playbook.items(), key=lambda item: int(item[0])):
        sql = section.get("sql", "")
        spec = catalog.get(number, {})
        question = spec.get("Câu hỏi") or section.get("title") or f"Air/OTA view #{number}"
        if not sql:
            guidance = strip_markdown(section.get("block", ""))
            sql = f"-- No standalone SQL template in playbook for view #{number}.\n-- Guidance: {guidance}"
        concepts = unique(
            [
                spec.get("Metric (đơn vị đếm)", ""),
                spec.get("Time column", ""),
                spec.get("Grain", ""),
                spec.get("Spec", ""),
                *re.findall(r"\b[A-Z]{2,6}\b", question),
            ]
        )
        explanation_parts = [
            f"View #{number}.",
            f"Make clear: {section['make_clear']}" if section.get("make_clear") else "",
            f"Spec: {spec.get('Spec')}" if spec.get("Spec") else "",
            f"Note: {spec.get('Note') or spec.get('Filter/Note')}" if spec.get("Note") or spec.get("Filter/Note") else "",
        ]
        example_id = f"qex_zlp_air_view_{int(number):02d}"
        examples[example_id] = {
            "id": example_id,
            "question": question,
            "sql": sql,
            "explanation": normalize_text(" ".join(part for part in explanation_parts if part)),
            "concepts": concepts,
            "used_tables": extract_tables(sql),
            "owner": OWNER,
            "status": "approved",
            "created_at": timestamp,
            "updated_at": timestamp,
        }
    return examples


def load_workbook_or_fail(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Missing dependency openpyxl. Run `venv/bin/pip install -r requirements.txt`.") from exc
    return load_workbook(path, read_only=True, data_only=True)


def header_map(row: tuple[Any, ...]) -> dict[str, int]:
    return {normalize_text(value).casefold(): index for index, value in enumerate(row) if normalize_text(value)}


def get_cell(row: tuple[Any, ...], headers: dict[str, int], *names: str) -> str:
    for name in names:
        index = headers.get(name.casefold())
        if index is not None and index < len(row):
            return normalize_text(row[index])
    return ""


def parse_dictionary(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook_or_fail(path)
    timestamp = now_iso()
    relationships: list[dict[str, str]] = []
    if "quan_he_2_bang" in workbook.sheetnames:
        sheet = workbook["quan_he_2_bang"]
        for row in sheet.iter_rows(min_row=4, values_only=True):
            left, right, note = (list(row) + ["", "", ""])[:3]
            left_text = normalize_text(left)
            right_text = normalize_text(right)
            if left_text and right_text:
                relationships.append(
                    {
                        "from": f"search_air.{left_text}",
                        "to": f"payment_air.{right_text}",
                        "type": normalize_text(note) or "mapping",
                    }
                )

    records: dict[str, dict[str, Any]] = {}
    for table in ["search_air", "payment_air"]:
        if table not in workbook.sheetnames:
            continue
        sheet = workbook[table]
        rows = sheet.iter_rows(values_only=True)
        headers = header_map(next(rows))
        columns = []
        for row in rows:
            name = get_cell(row, headers, "Column")
            if not name:
                continue
            data_type = get_cell(row, headers, "Data type")
            meaning = get_cell(row, headers, "Ý nghĩa", "Ý nghĩa (Trang điền)")
            usage = get_cell(row, headers, "Cách dùng", "Cách dùng (Trang điền)")
            format_notes = get_cell(row, headers, "Format notes & QA flags")
            examples = get_cell(row, headers, "Example values")
            columns.append(
                {
                    "name": name,
                    "business_meaning": normalize_text(". ".join(part for part in [meaning, usage, format_notes] if part)),
                    "data_type": data_type,
                    "aliases": unique([meaning, usage, examples]),
                }
            )
        record_id = f"dict_zlp_air_{table}"
        records[record_id] = {
            "id": record_id,
            "table": table,
            "description": f"Zalopay AIR/OTA data dictionary for `{table}`.",
            "columns": columns,
            "relationships": relationships,
            "owner": OWNER,
            "status": "approved",
            "created_at": timestamp,
            "updated_at": timestamp,
        }
    return records


def build_payloads() -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    knowledge = {
        "schema_version": 1,
        "knowledge": parse_metric_definitions(ZALOPAY_DIR / "ref_zlp_metric_definitions.md"),
    }
    dictionary = {
        "schema_version": 1,
        "records": parse_dictionary(ZALOPAY_DIR / "data_dictionary_air.xlsx"),
    }
    examples = {
        "schema_version": 1,
        "examples": parse_question_examples(
            ZALOPAY_DIR / "catalog_question_bank_air.md",
            ZALOPAY_DIR / "playbook_sql_air_views.md",
        ),
    }
    return knowledge, dictionary, examples


def merge_json_payloads(
    knowledge: dict[str, Any],
    dictionary: dict[str, Any],
    examples: dict[str, Any],
) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    existing_knowledge = load_json(DATA_DIR / "knowledge_base.json", {"schema_version": 1, "knowledge": {}})
    existing_dictionary = load_json(DATA_DIR / "data_dictionary.json", {"schema_version": 1, "records": {}})
    existing_examples = load_json(DATA_DIR / "question_examples.json", {"schema_version": 1, "examples": {}})

    existing_knowledge.setdefault("knowledge", {}).update(knowledge["knowledge"])
    existing_dictionary.setdefault("records", {}).update(dictionary["records"])
    existing_examples.setdefault("examples", {}).update(examples["examples"])
    return existing_knowledge, existing_dictionary, existing_examples


def push_sql(knowledge: dict[str, Any], dictionary: dict[str, Any], examples: dict[str, Any]) -> None:
    if load_dotenv:
        load_dotenv(ROOT / ".env")
    database_url = os.getenv("DATABASE_URL", "")
    if not database_url:
        raise RuntimeError("DATABASE_URL is required for --push-sql")
    sys.path.insert(0, str(ROOT))
    from knowledge_store import PostgresStorage

    storage = PostgresStorage(database_url)
    storage.bootstrap()
    storage.save_knowledge_base(knowledge)
    storage.save_data_dictionary(dictionary)
    storage.save_question_examples(examples)


def print_summary(knowledge: dict[str, Any], dictionary: dict[str, Any], examples: dict[str, Any]) -> None:
    print("Zalopay AIR ingest preview")
    print(f"- knowledge records: {len(knowledge['knowledge'])}")
    print(f"- dictionary records: {len(dictionary['records'])}")
    print(f"- question examples: {len(examples['examples'])}")
    sample_knowledge = next(iter(knowledge["knowledge"].values()), {})
    sample_dictionary = next(iter(dictionary["records"].values()), {})
    sample_example = next(iter(examples["examples"].values()), {})
    print(f"- sample knowledge: {sample_knowledge.get('id')} / {sample_knowledge.get('name')}")
    print(f"- sample dictionary: {sample_dictionary.get('id')} / {len(sample_dictionary.get('columns', []))} columns")
    print(f"- sample example: {sample_example.get('id')} / {sample_example.get('question')}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Ingest Zalopay AIR references into local JSON or Postgres.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print counts without writing.")
    parser.add_argument("--write-json", action="store_true", help="Merge parsed records into data/*.json.")
    parser.add_argument("--push-sql", action="store_true", help="Push merged payloads to Postgres using DATABASE_URL.")
    args = parser.parse_args()

    if not (args.dry_run or args.write_json or args.push_sql):
        args.dry_run = True

    knowledge, dictionary, examples = build_payloads()
    merged = merge_json_payloads(knowledge, dictionary, examples)
    print_summary(knowledge, dictionary, examples)

    if args.write_json:
        save_json(DATA_DIR / "knowledge_base.json", merged[0])
        save_json(DATA_DIR / "data_dictionary.json", merged[1])
        save_json(DATA_DIR / "question_examples.json", merged[2])
        print("- wrote JSON: data/knowledge_base.json, data/data_dictionary.json, data/question_examples.json")
    if args.push_sql:
        push_sql(*merged)
        print("- pushed to Postgres")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
